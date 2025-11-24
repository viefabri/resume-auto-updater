import json
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries
import os
import datetime
import copy
import config
import sys

# Configuration for this script
MASTER_JSON_PATH = config.OUTPUT_FILE
UPDATE_JSON_PATH = os.path.join('005_ToolOutput', '02_ResumeUpdate', 'Data', 'resume_update.json')
TEMPLATE_EXCEL_PATH = '経歴書_Updated_20251124.xlsx'
TEMPLATE_SHEET_NAME = '_Template'
TARGET_SHEET_NAME = 'スキルシート'
START_ROW = 21

def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def merge_data(master_data, update_data):
    payloads = update_data.get('update_payload', [])
    insert_count = 0
    
    # Process all payloads
    for payload in payloads:
        action = payload.get('action')
        target_no = payload.get('target_no')
        new_data = payload.get('data')

        if action == 'INSERT' and target_no == 0:
            # Insert at the beginning
            master_data['work_history'].insert(0, new_data)
            insert_count += 1
        elif action == 'UPDATE':
            # Find and update
            found = False
            for i, entry in enumerate(master_data['work_history']):
                if entry.get('no') == str(target_no):
                    master_data['work_history'][i] = new_data
                    found = True
                    break
            if not found:
                print(f"Warning: Entry with no {target_no} not found for UPDATE.")
        else:
            print(f"Warning: Unknown action {action} or target_no {target_no}")

    # Renumber
    for i, entry in enumerate(master_data['work_history']):
        entry['no'] = str(i + 1)
        
    # Update Footer if present
    footer_update = update_data.get('footer_update')
    if footer_update and footer_update.get('update_required'):
        master_data['footer']['other_col_b'] = footer_update.get('other_col_b', [])

    return master_data

def clean_sheet(ws):
    """Clears all content and merged cells from START_ROW onwards."""
    # Remove any merged cells that intersect with or are below START_ROW
    to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_row >= START_ROW:
            to_unmerge.append(merged_range)
            
    print(f"Ranges to unmerge: {[str(r) for r in to_unmerge]}")
    for mr in to_unmerge:
        try:
            # print(f"Removing merged range: {mr}")
            ws.unmerge_cells(str(mr))
        except KeyError:
            print(f"Warning: Could not unmerge {mr} via unmerge_cells. Forcing removal from ranges.")
            if mr in ws.merged_cells.ranges:
                ws.merged_cells.remove(mr)
        except Exception as e:
            print(f"Warning: Error unmerging {mr}: {e}")

    max_row = ws.max_row
    if max_row >= START_ROW:
        print(f"Deleting rows {START_ROW} to {max_row}...")
        ws.delete_rows(START_ROW, amount=max_row - START_ROW + 1)

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = copy.copy(src_cell.number_format)
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)

import logging
logging.basicConfig(filename='debug.log', level=logging.DEBUG, filemode='w', format='%(asctime)s - %(levelname)s - %(message)s')

def ensure_writable(ws, row, col):
    cell = ws.cell(row=row, column=col)
    
    if type(cell).__name__ == 'MergedCell':
        logging.debug(f"Found MergedCell at {cell.coordinate}. Attempting to unmerge.")
        found = False
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                found = True
                try:
                    ws.unmerge_cells(str(mr))
                except:
                    if mr in ws.merged_cells.ranges:
                        ws.merged_cells.remove(mr)
                break
        
        if not found:
            logging.error(f"PANIC: Cell {cell.coordinate} is MergedCell but NOT found in any range!")
            
        # Force re-fetch
        cell = ws.cell(row=row, column=col)
        
        if type(cell).__name__ == 'MergedCell':
             logging.error(f"ERROR: ensure_writable failed for {cell.coordinate}. Still MergedCell. Force-Fixing...")
             try:
                 # Force replace the cell in the worksheet's internal storage
                 from openpyxl.cell.cell import Cell
                 new_cell = Cell(ws, row=row, column=col)
                 if hasattr(ws, '_cells'):
                     ws._cells[(row, col)] = new_cell
                     logging.debug(f"Force-fixed {cell.coordinate} via _cells")
                 else:
                     logging.error("Worksheet has no _cells attribute. Cannot force-fix.")
                 cell = new_cell
             except Exception as e:
                 logging.error(f"Failed to force-fix {cell.coordinate}: {e}")

        else:
             logging.debug(f"Fixed {cell.coordinate}. New type: {type(cell)}")
    
    logging.debug(f"Returning {cell.coordinate} as {type(cell)}")
    return cell

def apply_template_and_write_data(ws, template_ws, entry, start_row):
    """Applies template styles and writes data for a single entry (5 rows)."""
    
    # 1. Apply Template (Styles and Merged Cells)
    # Template is rows 1-5 in _Template sheet
    for row_idx in range(1, 6):
        for col_idx in range(1, 32): # A to AE (Index 1 to 31)
            src_cell = template_ws.cell(row=row_idx, column=col_idx)
            dst_cell = ws.cell(row=start_row + row_idx - 1, column=col_idx)
            copy_style(src_cell, dst_cell)
            # We don't copy value here, we will write data next

    # Copy merged cells from template to current block
    for merged_range in template_ws.merged_cells.ranges:
        # merged_range is like 'A1:A5'
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        
        # Shift row to current target
        new_min_row = min_row + start_row - 1
        new_max_row = max_row + start_row - 1
        
        # Merge in target
        ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)

    # 2. Write Data
    
    # No (A)
    ensure_writable(ws, start_row, 1).value = entry.get('no')
    
    # Period (B)
    ensure_writable(ws, start_row, 2).value = entry.get('period', {}).get('start')
    
    # End: B3 -> B{start_row+2}
    try:
        ensure_writable(ws, start_row + 2, 2).value = entry.get('period', {}).get('end')
    except AttributeError:
        print(f"CRITICAL: Failed to write to B{start_row+2}. It is MergedCell.")
        cell = ws.cell(row=start_row + 2, column=2)
        print(f"Cell type: {type(cell)}")
        print(f"Ranges containing B{start_row+2}: {[str(r) for r in ws.merged_cells.ranges if cell.coordinate in r]}")
        sys.stdout.flush()
        raise
    
    # Business Content
    bc = entry.get('business_content', {})
    titles = bc.get('title_col_e', [])
    roles = bc.get('role_col_f', [])
    details = bc.get('detail_col_g', [])
    
    print(f"DEBUG: titles len={len(titles)}, roles len={len(roles)}, details len={len(details)}")
    
    for i in range(5):
        print(f"DEBUG: Loop i={i}")
        try:
            if i < len(titles): 
                # print(f"Writing title at {start_row+i}, 5")
                ensure_writable(ws, start_row + i, 5).value = titles[i]
            if i < len(roles): 
                # print(f"Writing role at {start_row+i}, 6")
                ensure_writable(ws, start_row + i, 6).value = roles[i]
            if i < len(details): 
                # print(f"Writing detail at {start_row+i}, 7")
                ensure_writable(ws, start_row + i, 7).value = details[i]
        except Exception as e:
            print(f"CRITICAL: Failed to write to row {start_row+i}. Error: {e}")
            sys.stdout.flush()
            raise
            print(f"CRITICAL: Failed to write to row {start_row+i}. MergedCell error.")
            sys.stdout.flush()
            raise
        
    # Technology
    tech = entry.get('technology', {})
    envs = tech.get('environment_col_u', [])
    langs = tech.get('language_col_z', [])
    procs = tech.get('process_col_ae', [])
    
    for i in range(5):
        if i < len(envs): ensure_writable(ws, start_row + i, 21).value = envs[i]
        if i < len(langs): ensure_writable(ws, start_row + i, 26).value = langs[i]
        if i < len(procs): ensure_writable(ws, start_row + i, 31).value = procs[i]

def write_footer(ws, footer_data, start_row):
    # Write footer starting at start_row
    others = footer_data.get('other_col_b', [])
    
    ws.cell(row=start_row, column=1).value = "その他"
    
    for i in range(len(others)):
        ws.cell(row=start_row + i, column=2).value = others[i]

def draw_border(ws, start_row, end_row):
    # Apply medium border to the outer edge of the range A{start_row}:AE{end_row}
    medium = Side(border_style="medium", color="000000")
    
    # Top edge (Row start_row)
    for col in range(1, 32): # A to AE
        cell = ws.cell(row=start_row, column=col)
        # If merged, we need to find the top-left to modify style
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             # Find range
             for mr in ws.merged_cells.ranges:
                 if cell.coordinate in mr:
                     cell = ws.cell(row=mr.min_row, column=mr.min_col)
                     break
        
        new_border = copy.copy(cell.border)
        new_border.top = medium
        cell.border = new_border
        
    # Bottom edge (Row end_row)
    for col in range(1, 32):
        cell = ws.cell(row=end_row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             for mr in ws.merged_cells.ranges:
                 if cell.coordinate in mr:
                     cell = ws.cell(row=mr.min_row, column=mr.min_col)
                     break
        
        new_border = copy.copy(cell.border)
        new_border.bottom = medium
        cell.border = new_border
        
    # Left edge (Column A)
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=1)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             for mr in ws.merged_cells.ranges:
                 if cell.coordinate in mr:
                     cell = ws.cell(row=mr.min_row, column=mr.min_col)
                     break
        
        new_border = copy.copy(cell.border)
        new_border.left = medium
        cell.border = new_border
        
    # Right edge (Column AE -> 31)
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=31)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             for mr in ws.merged_cells.ranges:
                 if cell.coordinate in mr:
                     cell = ws.cell(row=mr.min_row, column=mr.min_col)
                     break
        
        new_border = copy.copy(cell.border)
        new_border.right = medium
        cell.border = new_border

def main():
    # 1. Load Data
    try:
        master_data = load_json(MASTER_JSON_PATH)
        update_data = load_json(UPDATE_JSON_PATH)
    except FileNotFoundError as e:
        print(f"Error loading JSON: {e}")
        return

    # 2. Merge
    master_data = merge_data(master_data, update_data)
    save_json(MASTER_JSON_PATH, master_data)
    print(f"Updated {MASTER_JSON_PATH}")

    # 3. Excel Manipulation
    try:
        wb = openpyxl.load_workbook(TEMPLATE_EXCEL_PATH)
    except FileNotFoundError:
        print(f"Error loading Excel: {TEMPLATE_EXCEL_PATH}")
        return
        
    if TARGET_SHEET_NAME not in wb.sheetnames or TEMPLATE_SHEET_NAME not in wb.sheetnames:
        print(f"Sheet {TARGET_SHEET_NAME} or {TEMPLATE_SHEET_NAME} not found.")
        return
        
    ws = wb[TARGET_SHEET_NAME]
    template_ws = wb[TEMPLATE_SHEET_NAME]
    
    # 4. Clean Sheet
    clean_sheet(ws)
    
    # 5. Render History
    current_row = START_ROW
    print("Rendering work history...")
    for entry in master_data['work_history']:
        print(f"Processing entry {entry.get('no')}")
        apply_template_and_write_data(ws, template_ws, entry, current_row)
        current_row += 5
        
    # 6. Render Footer
    print("Writing footer...")
    write_footer(ws, master_data['footer'], current_row)
    
    # 7. Draw Border
    # Border covers only the history part (START_ROW to current_row - 1)
    print("Drawing border...")
    draw_border(ws, START_ROW, current_row - 1)
    
    # 8. Save
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    output_filename = f"経歴書_Updated_{timestamp}.xlsx"
    wb.save(output_filename)
    print(f"Saved {output_filename}")

if __name__ == "__main__":
    main()
